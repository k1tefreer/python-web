from django.shortcuts import render, redirect, HttpResponse
from app01 import models
from app01.utils.pagination import Pagination
from app01.utils.form import UserModelForm,PrettyEditModleForm,PrettyModleForm

# views可以拆分 models.py 一定不能拆分

def depart_list(request):
    ''' 部门列表 '''


    #  不需要了 已经有中间件了 再middleware里的auth
    # info = request.session.get('info')
    # if not info:
    #     return redirect('/login/')

    # 1.去数据库中获取所有的部门列表
    # queryset 意为 [列表,列表,列表]
    queryset = models.Department.objects.all()
    page_object = Pagination(request,queryset,page_size=5)
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    ''' 添加部门 '''


    # info = request.session.get('info')
    # if not info:
    #     return redirect('/login/')


    if request.method == 'GET':
        return render(request, 'depart_add.html')

    # 获取用户POST提交过来的数据 (title输入为空)
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重定向回部门列表
    return redirect("/depart/list")


def depart_delete(request):
    ''' 删除部门 '''



    # 获取ID
    # http://127.0.0.1:8000/depart/delete/?nid=1
    nid = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=nid).delete()

    # 重定向回部门列表
    return redirect("/depart/list/")

    # http://127.0.0.1:8000/depart/10/edit/ 相当于正则表达式
    # http://127.0.0.1:8000/depart/2/edit/


def depart_edit(request, nid):
    '''修改部门'''


    if request.method == "GET":
        # 根据nid, 获取它的数据 [obj,]
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id,row_object.title)
        return render(request, 'depart_edit.html', {"row_object": row_object})

    # 获取用户提交的数据
    title = request.POST.get("title")

    # 根据ID找到数据库中的数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_multi(request):
    """ 批量上传 （Excel文件） """
    from django.core.files.uploadedfile import InMemoryUploadedFile
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    print(type(file_object))

    # 2.对象传递给openpyxl, 由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

            # # 获取第一行第一列
            # cell = sheet.cell(1, 1)
            # print(cell.value)

    # 3.循环获取每一行每一列
    for row in sheet.iter_rows(min_row=1):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            # 把获取的数据上传到数据库
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')